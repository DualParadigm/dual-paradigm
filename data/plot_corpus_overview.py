from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


WORKBOOK_PATH = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/新维度语料库0310_fixed (1).xlsx"
)
OVERVIEW_OUTPUT = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/corpus_overview_updated.png"
)
FIELD_OUTPUT = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_field_distribution.png"
)


# Keep the original palette logic, only relabeling categories to fit the current sheet.
VENUE_COLORS = {
    "Computer science": "#a8c8d3",
    "HCI": "#728CB7",
    "Sociology": "#61B3E9",
    "VIS": "#8E568F",
    "other": "#FF7685",
}

TYPE_COLORS = {
    "application+system": "#71629A",
    "technique+algorithm": "#67E8C8",
    "theory": "#C85ABB",
}

VENUE_LABELS = {
    "Computer science": "Computer science",
    "HCI": "HCI-related",
    "Sociology": "Sociology",
    "VIS": "VIS-related",
    "other": "Other",
}

TYPE_LABELS = {
    "application+system": "Application & System",
    "technique+algorithm": "Technique & Algorithm",
    "theory": "Theoretical & Empirical",
}


def load_paper_rows():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["paper"]
    rows = []
    for r in range(3, ws.max_row + 1):
        kind = ws.cell(r, 5).value
        year = ws.cell(r, 8).value
        source = ws.cell(r, 1).value
        if kind != "论文" or not year or not source:
            continue
        rows.append(
            {
                "year": int(year),
                "venue_cat": (ws.cell(r, 3).value or "other").strip(),
                "venue": ws.cell(r, 4).value,
                "paper_type": (ws.cell(r, 6).value or "theory").strip(),
            }
        )
    return rows


def stacked_counts(rows, key):
    years = sorted({row["year"] for row in rows})
    counts = defaultdict(Counter)
    for row in rows:
        counts[row["year"]][row[key]] += 1
    return years, counts


def style_axis(ax, ymax):
    ax.set_ylabel("Number of Papers", fontsize=16)
    ax.set_ylim(0, ymax)
    ax.set_yticks(np.arange(0, ymax + 0.1, 3))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#333333")
    ax.spines["bottom"].set_color("#333333")
    ax.spines["left"].set_linewidth(1.0)
    ax.spines["bottom"].set_linewidth(1.0)
    ax.tick_params(axis="x", rotation=50, labelsize=9)
    ax.tick_params(axis="y", labelsize=10)


def plot_overview(rows):
    venue_order = ["Computer science", "HCI", "Sociology", "VIS", "other"]
    type_order = ["application+system", "technique+algorithm", "theory"]

    years, venue_counts = stacked_counts(rows, "venue_cat")
    _, type_counts = stacked_counts(rows, "paper_type")

    yearly_totals = Counter(row["year"] for row in rows)
    ymax = int(np.ceil((max(yearly_totals.values()) + 1) / 3.0) * 3)

    plt.rcParams["font.family"] = "DejaVu Sans"
    fig, axes = plt.subplots(2, 1, figsize=(8, 8))

    x = np.arange(len(years))
    width = 0.82

    bottom = np.zeros(len(years))
    for category in venue_order:
        values = [venue_counts[year].get(category, 0) for year in years]
        axes[0].bar(
            x,
            values,
            width=width,
            bottom=bottom,
            color=VENUE_COLORS[category],
            edgecolor="none",
            label=VENUE_LABELS[category],
        )
        bottom += np.array(values)
    axes[0].set_title(
        "(a) Distribution of paper venues over time",
        fontsize=18,
        fontweight="bold",
        y=0.95,
    )
    axes[0].legend(frameon=False, loc="upper left", fontsize=12, bbox_to_anchor=(0.0, 0.9))
    axes[0].set_xticks(x, years)
    style_axis(axes[0], ymax)

    bottom = np.zeros(len(years))
    for paper_type in type_order:
        values = [type_counts[year].get(paper_type, 0) for year in years]
        axes[1].bar(
            x,
            values,
            width=width,
            bottom=bottom,
            color=TYPE_COLORS[paper_type],
            edgecolor="none",
            label=TYPE_LABELS[paper_type],
        )
        bottom += np.array(values)
    axes[1].set_title(
        "(b) Distribution of paper types over time",
        fontsize=18,
        fontweight="bold",
        y=0.95,
    )
    axes[1].legend(frameon=False, loc="upper left", fontsize=12, bbox_to_anchor=(0.0, 0.86))
    axes[1].set_xticks(x, years)
    style_axis(axes[1], ymax)

    fig.tight_layout(h_pad=2.0)
    fig.savefig(OVERVIEW_OUTPUT, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_field_distribution(rows):
    venue_order = ["Computer science", "HCI", "Sociology", "VIS", "other"]
    counts = Counter(row["venue_cat"] for row in rows)
    labels = [VENUE_LABELS[k] for k in venue_order]
    values = [counts.get(k, 0) for k in venue_order]
    colors = [VENUE_COLORS[k] for k in venue_order]

    plt.rcParams["font.family"] = "DejaVu Sans"
    fig, ax = plt.subplots(figsize=(8, 4.6))
    y = np.arange(len(labels))
    bars = ax.barh(y, values, color=colors, edgecolor="none", height=0.72)
    ax.set_yticks(y, labels)
    ax.invert_yaxis()
    ax.set_xlabel("Number of Papers", fontsize=15)
    ax.set_title("Distribution of paper fields", fontsize=18, fontweight="bold", pad=14)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["bottom"].set_color("#333333")
    ax.tick_params(axis="both", labelsize=12)
    ax.set_xlim(0, max(values) + 6)
    for bar, value in zip(bars, values):
        ax.text(value + 0.4, bar.get_y() + bar.get_height() / 2, str(value), va="center", fontsize=11)
    fig.tight_layout()
    fig.savefig(FIELD_OUTPUT, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def main():
    rows = load_paper_rows()
    plot_overview(rows)
    plot_field_distribution(rows)
    print(f"overview={OVERVIEW_OUTPUT}")
    print(f"field={FIELD_OUTPUT}")
    print(f"papers={len(rows)}")


if __name__ == "__main__":
    main()
