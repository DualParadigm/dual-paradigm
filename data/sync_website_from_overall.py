from __future__ import annotations

import json
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/primav/Documents/GitHub/dual-paradigm")
WORKBOOK_PATH = ROOT / "data" / "新维度语料库0310_fixed (1).xlsx"
IMAGE_ZIP_PATH = ROOT / "data" / "267.zip"
DATA_JSON_PATH = ROOT / "list" / "data.json"
OLD_DATA_JSON_PATH = ROOT / "list" / "data.before-sync.bak.json"
THUMB_DIR = ROOT / "thumbnail"


GROUP_KEY_MAP = {
    "User (who)": "User",
    "Topic (what)": "Topic",
    "Presentation (how)": "Presentation",
    "Goal (why)": "Goal",
    "data origin (when) historical, live, predictive, synthesized": "dataOrigin",
    "representation form (how)": "representationForm",
    "Scale": "SituatednessScale",
    "Semantics": "SituatednessSemantics",
}

VALUE_MAP = {
    "Individual": "Individual",
    "Group": "Group",
    "Public": "Public",
    "Entertainment, Sports &Culture": "Entertainment, Sports&Culture",
    "Environment (Surroudings/Nature), Geography&Ecology": "Environment (Surroudings/Nature), Geography&Ecology",
    "Health, Life & Emotion": "Health, Life & Emotion",
    "Science": "Science",
    "Society&History (include media)": "Society&History(include media)",
    "Industry, Architecture&Engineering": "Industry, Architecture&Engineering",
    "Unspecified&Mixed Topics": "Unspecified&Mixed Topics",
    "XR": "XR",
    "Event": "Event",
    "Artifact": "Artifact",
    "Installation": "Installation",
    "comprehension": "comprehension",
    "communication": "communication",
    "archive": "archive",
    "engagement": "engagement",
    "reflection": "reflection",
    "analysis": "analysis",
    "Environmental data": "Environmental data",
    "Participant-generated": "Participant-generated",
    "Collected&Uploaded": "Collected&Uploaded",
    "Unspecified": "Unspecified",
    "visual": "visual",
    "haptic": "haptic",
    "sound": "sound",
    "other (taste, smell，action)": "other (taste, smell)",
    "small": "small",
    "medium": "medium",
    "large": "large",
    "Symbolic": "Symbolic",
    "Iconic": "Iconic",
    "Indexical": "Indexical",
}


MANUAL_TITLE_MAP = {
    "https://anatomage.com/table/": "Anatomage Table",
    "https://www.instagram.com/reel/C8rlc5nJtV0/?epik=dj0yJnU9UkctWXBiZHlqcWNraU9wNHA4SWZrdmxkeDlYZHNfUDAmcD0wJm49d3prNDJEYzdyazQ4dUlYUks0LWQtQSZ0PUFBQUFBR2lZTER3": "Interactive LED Screens for Museum",
    "https://interactiveparty.com/product/multiball-led-basketball/": "Multiball LED Basketball",
    "https://www.youtube.com/shorts/06PhXPWQa9k": "Breathing Digital Life Into Cultural Objects",
    "https://www.youtube.com/watch?v=44pqxnMsW-0": "WebAR Brochure to Interactive 3D Model",
    "https://www.legoland.com/florida/things-to-do/theme-park/rides-attractions/imagination-zone/": "Imagination Zone Interactive Digital Art Installation",
    "https://chalkacademy.com/montessori-multiplication-division-boards/": "Montessori Multiplication and Division Boards",
    "https://mye.experimenta.science/exponate/uebersicht": "Light Buckets at Experimenta",
    "https://mye.experimenta.science/exponate/uebersicht/193b2ad0-c407-421d-b972-acba2252bcb9": "Packaging Signals Station at Experimenta",
}

MANUAL_SOURCE_BY_ID = {
    23: "http://dataphys.org/list/dan-gilberts-tv-ads/",
    24: "http://dataphys.org/list/dan-gilberts-tv-ads/",
}

MANUAL_TITLE_BY_ID = {
    23: "Dan Gilbert's TV Ads",
    24: "Dan Gilbert's TV Ads",
}


def normalize_text(s: str | None) -> str | None:
    if not s:
        return None
    return re.sub(r"\s+", " ", str(s)).strip()


def normalize_source(url: str | None) -> str | None:
    s = normalize_text(url)
    if not s:
        return None
    s = s.replace(" ", "")
    s = s.split("?")[0] if "instagram.com/accounts/login" not in s else s
    arnumber = re.search(r"(?:arnumber=|document/)(\d+)", s)
    if arnumber:
        return f"ieee:{arnumber.group(1)}"
    doiish = None
    for pat in [r"(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)", r"pii/([A-Z0-9]+)"]:
        m = re.search(pat, s, re.I)
        if m:
            doiish = m.group(1).rstrip(")/")
            break
    if doiish:
        return f"doi:{doiish.lower()}"
    return s.lower().rstrip("/")


def derive_title_from_url(url: str) -> str:
    cleaned = url.rstrip("/").split("/")[-1]
    if not cleaned:
        cleaned = url.rstrip("/").split("/")[-2]
    cleaned = cleaned.split("?")[0]
    cleaned = cleaned.replace("-", " ").replace("_", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned.title() if cleaned else "Untitled"


def normalize_year(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def build_title_maps():
    source_path = OLD_DATA_JSON_PATH if OLD_DATA_JSON_PATH.exists() else DATA_JSON_PATH
    existing = json.loads(source_path.read_text(encoding="utf-8"))
    existing_map = {
        normalize_source(item.get("Sourcelink")): item.get("Title")
        for item in existing
        if item.get("Sourcelink") and item.get("Title")
    }
    existing_by_id = {item.get("id"): item for item in existing if item.get("id") is not None}

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    paper_ws = wb["paper"]
    case_ws = wb["case"]
    paper_map = {}
    year_map = {}
    for r in range(3, paper_ws.max_row + 1):
        src = normalize_text(paper_ws.cell(r, 1).value)
        title = normalize_text(paper_ws.cell(r, 7).value)
        year = normalize_year(paper_ws.cell(r, 8).value)
        if src and title:
            paper_map[normalize_source(src)] = title
        if src and year is not None:
            year_map[normalize_source(src)] = year

    for r in range(3, case_ws.max_row + 1):
        src = normalize_text(case_ws.cell(r, 1).value)
        year = normalize_year(case_ws.cell(r, 3).value)
        if src and year is not None:
            year_map[normalize_source(src)] = year

    return existing_map, existing_by_id, paper_map, year_map


def parse_feature_name(feature_name: str):
    group, value = feature_name.split("__", 1)
    group = normalize_text(group)
    value = normalize_text(value)
    return GROUP_KEY_MAP[group], VALUE_MAP[value]


def build_records():
    existing_titles, existing_by_id, paper_titles, years_by_source = build_title_maps()
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["overall_matrix"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    feature_cols = []
    for c in range(5, ws.max_column + 1, 2):
        a_header = headers[c - 1]
        if str(a_header).endswith("__coderA"):
            feature_cols.append((c, str(a_header)[:-8]))

    records = []
    for r in range(2, ws.max_row + 1):
        source = normalize_text(ws.cell(r, 3).value)
        desc = normalize_text(ws.cell(r, 4).value) or ""
        case_id = int(ws.cell(r, 1).value)
        old_record = existing_by_id.get(case_id, {})
        if not source:
            source = normalize_text(old_record.get("Sourcelink"))
        if not source:
            source = MANUAL_SOURCE_BY_ID.get(case_id)
        if not source:
            continue

        source_key = normalize_source(source)
        title = (
            existing_titles.get(source_key)
            or old_record.get("Title")
            or MANUAL_TITLE_BY_ID.get(case_id)
            or paper_titles.get(source_key)
            or MANUAL_TITLE_MAP.get(source)
            or derive_title_from_url(source)
        )

        record = {
            "Title": title,
            "Sourcelink": source,
            "Description": desc,
            "User": [],
            "Topic": [],
            "Presentation": [],
            "Goal": [],
            "dataOrigin": [],
            "representationForm": [],
            "SituatednessScale": [],
            "SituatednessSemantics": [],
            "years": years_by_source.get(source_key, normalize_year(old_record.get("years"))),
            "id": case_id,
        }

        for c, feature_name in feature_cols:
            a_val = int(ws.cell(r, c).value or 0)
            b_val = int(ws.cell(r, c + 1).value or 0)
            if not (a_val or b_val):
                continue
            key, value = parse_feature_name(feature_name)
            record[key].append(value)

        records.append(record)

    return records


def sync_thumbnails():
    with zipfile.ZipFile(IMAGE_ZIP_PATH, "r") as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".png"):
                continue
            row_num = int(Path(name).stem)
            website_id = row_num - 2
            target = THUMB_DIR / f"{website_id}.png"
            with zf.open(name) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)


def main():
    backup_path = DATA_JSON_PATH.with_suffix(".before-sync.bak.json")
    shutil.copy2(DATA_JSON_PATH, backup_path)

    records = build_records()
    DATA_JSON_PATH.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    sync_thumbnails()

    print(f"backup={backup_path}")
    print(f"records={len(records)}")
    print(f"first_id={records[0]['id']}")
    print(f"last_id={records[-1]['id']}")


if __name__ == "__main__":
    main()
