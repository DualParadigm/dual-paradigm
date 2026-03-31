from __future__ import annotations

import csv
import math
import re
from collections import Counter
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from scipy.stats import chi2_contingency, fisher_exact
from sklearn.compose import ColumnTransformer
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegressionCV
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler


WORKBOOK_PATH = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/新维度语料库0310_fixed (1).xlsx"
)
MERGED_CSV = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_dimension_analysis_dataset.csv"
)
ASSOC_CSV = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_il_association_summary.csv"
)
GROUP_RATE_CSV = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_il_group_rates.csv"
)
BIVARIATE_CSV = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_overall_bivariate.csv"
)
L1_CSV = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_il_l1_logistic_coefficients.csv"
)
HEATMAP_OUTPUT = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_il_group_rate_heatmap.png"
)
TOP_OUTPUT = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/paper_overall_top_associations.png"
)


OUTCOME_MAP = {
    "I_design": 9,
    "J_presentation": 10,
    "K_representation": 11,
    "L_data": 12,
}


def normalize_value(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def binary_mark(v) -> int:
    return 0 if normalize_value(v) is None else 1


def extract_ieee_arnumber(url: str) -> str | None:
    match = re.search(r"(?:arnumber=|document/)(\d+)", url)
    return match.group(1) if match else None


def extract_doiish(url: str) -> str | None:
    patterns = [
        r"(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)",
        r"pii/([A-Z0-9]+)",
    ]
    for pat in patterns:
        match = re.search(pat, url, flags=re.I)
        if match:
            return match.group(1).rstrip(")/")
    return None


def normalize_source(url: str | None) -> str | None:
    if not url:
        return None
    s = str(url).strip()
    if not s:
        return None
    s = s.replace(" ", "")
    arnumber = extract_ieee_arnumber(s)
    if arnumber:
        return f"ieee:{arnumber}"
    doiish = extract_doiish(s)
    if doiish:
        return f"doi:{doiish.lower()}"
    s = s.lower()
    s = s.split("?")[0]
    s = s.rstrip("/")
    return s


def cramers_v(table: np.ndarray) -> float:
    chi2 = chi2_contingency(table, correction=False)[0]
    n = table.sum()
    if n == 0:
        return float("nan")
    r, k = table.shape
    denom = min(k - 1, r - 1)
    if denom <= 0:
        return float("nan")
    return math.sqrt((chi2 / n) / denom)


def phi_from_2x2(table: np.ndarray) -> float:
    chi2 = chi2_contingency(table, correction=False)[0]
    n = table.sum()
    if n == 0:
        return float("nan")
    return math.sqrt(chi2 / n)


def round_or_blank(value, digits=4):
    if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        return ""
    return f"{value:.{digits}f}"


def load_data():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    paper_ws = wb["paper"]
    matrix_ws = wb["overall_matrix"]

    paper_rows = []
    for r in range(3, paper_ws.max_row + 1):
        if paper_ws.cell(r, 5).value != "论文":
            continue
        src = normalize_value(paper_ws.cell(r, 1).value)
        if not src:
            continue
        row = {
            "paper_row": r,
            "source_link": src,
            "source_key": normalize_source(src),
            "venue_cat": normalize_value(paper_ws.cell(r, 3).value) or "Unknown",
            "venue": normalize_value(paper_ws.cell(r, 4).value) or "Unknown",
            "paper_type": normalize_value(paper_ws.cell(r, 6).value) or "Unknown",
            "title": normalize_value(paper_ws.cell(r, 7).value) or f"paper_row_{r}",
            "year": paper_ws.cell(r, 8).value,
        }
        for outcome, col in OUTCOME_MAP.items():
            row[outcome] = binary_mark(paper_ws.cell(r, col).value)
        paper_rows.append(row)
    paper_df = pd.DataFrame(paper_rows)

    matrix_headers = [matrix_ws.cell(1, c).value for c in range(1, matrix_ws.max_column + 1)]
    overall_rows = []
    for r in range(2, matrix_ws.max_row + 1):
        row = {
            "case_id": matrix_ws.cell(r, 1).value,
            "overall_row": matrix_ws.cell(r, 2).value,
            "source_link_overall": normalize_value(matrix_ws.cell(r, 3).value),
            "source_key": normalize_source(matrix_ws.cell(r, 3).value),
        }
        for c in range(5, matrix_ws.max_column + 1, 2):
            a_name = str(matrix_headers[c - 1])
            b_name = str(matrix_headers[c])
            feature = a_name.replace("__coderA", "")
            a_val = int(matrix_ws.cell(r, c).value or 0)
            b_val = int(matrix_ws.cell(r, c + 1).value or 0)
            row[f"{feature}__union"] = int(a_val or b_val)
            row[f"{feature}__intersection"] = int(a_val and b_val)
        overall_rows.append(row)
    overall_df = pd.DataFrame(overall_rows)
    return paper_df, overall_df


def build_merged():
    paper_df, overall_df = load_data()
    merged = paper_df.merge(overall_df, how="left", on="source_key")
    merged["matched_overall"] = merged["overall_row"].notna().astype(int)
    return merged


def association_tests(paper_df: pd.DataFrame):
    rows = []
    for predictor in ["venue_cat", "paper_type"]:
        for outcome in OUTCOME_MAP:
            table = pd.crosstab(paper_df[predictor], paper_df[outcome])
            chi2, p_value, _, _ = chi2_contingency(table, correction=False)
            rows.append(
                {
                    "predictor": predictor,
                    "outcome": outcome,
                    "n": int(table.to_numpy().sum()),
                    "levels": int(table.shape[0]),
                    "chi2": round_or_blank(chi2),
                    "p_value": round_or_blank(p_value),
                    "cramers_v": round_or_blank(cramers_v(table.to_numpy())),
                }
            )
    return pd.DataFrame(rows)


def group_rates(paper_df: pd.DataFrame):
    rows = []
    for predictor in ["venue_cat", "paper_type"]:
        for group_value, subdf in paper_df.groupby(predictor):
            for outcome in OUTCOME_MAP:
                rows.append(
                    {
                        "predictor": predictor,
                        "group": group_value,
                        "outcome": outcome,
                        "n": len(subdf),
                        "positive_n": int(subdf[outcome].sum()),
                        "positive_rate": subdf[outcome].mean(),
                    }
                )
    return pd.DataFrame(rows)


def bivariate_overall_tests(merged: pd.DataFrame):
    feature_cols = [
        c for c in merged.columns if c.endswith("__union") and "__coder" not in c
    ]
    matched = merged[merged["matched_overall"] == 1].copy()
    rows = []
    for outcome in OUTCOME_MAP:
        for feature in feature_cols:
            table = pd.crosstab(matched[feature], matched[outcome]).reindex(
                index=[0, 1], columns=[0, 1], fill_value=0
            )
            odds_ratio, fisher_p = fisher_exact(table.to_numpy())
            rows.append(
                {
                    "outcome": outcome,
                    "overall_feature": feature.replace("__union", ""),
                    "n": int(table.to_numpy().sum()),
                    "feature_positive_n": int(table.loc[1].sum()),
                    "outcome_positive_n": int(table[1].sum()),
                    "both_positive_n": int(table.loc[1, 1]),
                    "odds_ratio": round_or_blank(odds_ratio),
                    "fisher_p": round_or_blank(fisher_p),
                    "phi": round_or_blank(phi_from_2x2(table.to_numpy())),
                }
            )
    return pd.DataFrame(rows)


def l1_logistic_models(merged: pd.DataFrame):
    matched = merged[merged["matched_overall"] == 1].copy()
    feature_cols = [
        c for c in matched.columns if c.endswith("__union") and "__coder" not in c
    ]
    base_numeric = ["year"]
    base_categorical = ["venue_cat", "paper_type"]

    result_rows = []
    for outcome in OUTCOME_MAP:
        y = matched[outcome].astype(int)
        X = matched[base_numeric + base_categorical + feature_cols].copy()

        preprocessor = ColumnTransformer(
            transformers=[
                (
                    "num",
                    Pipeline(
                        steps=[
                            ("imputer", SimpleImputer(strategy="median")),
                            ("scaler", StandardScaler()),
                        ]
                    ),
                    base_numeric,
                ),
                (
                    "cat",
                    OneHotEncoder(drop="first", handle_unknown="ignore"),
                    base_categorical,
                ),
                ("bin", "passthrough", feature_cols),
            ]
        )

        model = LogisticRegressionCV(
            Cs=10,
            cv=5,
            penalty="l1",
            solver="liblinear",
            max_iter=10000,
            scoring="roc_auc",
            class_weight="balanced",
            random_state=42,
        )

        pipe = Pipeline([("prep", preprocessor), ("model", model)])
        pipe.fit(X, y)

        prep = pipe.named_steps["prep"]
        feature_names = prep.get_feature_names_out()
        coefs = pipe.named_steps["model"].coef_[0]

        for fname, coef in zip(feature_names, coefs):
            if abs(coef) < 1e-8:
                continue
            result_rows.append(
                {
                    "outcome": outcome,
                    "feature": fname,
                    "coefficient": coef,
                    "odds_ratio": math.exp(coef),
                    "direction": "positive" if coef > 0 else "negative",
                }
            )

    return pd.DataFrame(result_rows)


def plot_group_rate_heatmap(group_rates_df: pd.DataFrame):
    plot_df = group_rates_df.copy()
    plot_df["row_name"] = plot_df["predictor"] + ": " + plot_df["group"]
    pivot = plot_df.pivot(index="row_name", columns="outcome", values="positive_rate")

    plt.rcParams["font.family"] = "DejaVu Sans"
    fig, ax = plt.subplots(figsize=(7.8, 6.8))
    sns.heatmap(
        pivot,
        annot=True,
        fmt=".2f",
        cmap="YlGnBu",
        linewidths=0.5,
        cbar_kws={"label": "Positive Rate"},
        ax=ax,
    )
    ax.set_title("Paper I-L outcome rates by venue and paper type", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("")
    ax.set_ylabel("")
    fig.tight_layout()
    fig.savefig(HEATMAP_OUTPUT, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_top_associations(bivar_df: pd.DataFrame):
    rows = []
    for outcome in OUTCOME_MAP:
        sub = bivar_df[bivar_df["outcome"] == outcome].copy()
        sub["abs_phi"] = sub["phi"].astype(float).abs()
        top = sub.sort_values("abs_phi", ascending=False).head(8)
        for _, rec in top.iterrows():
            rows.append(
                {
                    "outcome": outcome,
                    "overall_feature": rec["overall_feature"],
                    "phi": float(rec["phi"]),
                }
            )
    plot_df = pd.DataFrame(rows)
    pivot = plot_df.pivot(index="overall_feature", columns="outcome", values="phi").fillna(0)

    plt.rcParams["font.family"] = "DejaVu Sans"
    fig, ax = plt.subplots(figsize=(8.5, 7.8))
    sns.heatmap(
        pivot,
        annot=True,
        fmt=".2f",
        cmap="coolwarm",
        center=0,
        linewidths=0.5,
        cbar_kws={"label": "Phi coefficient"},
        ax=ax,
    )
    ax.set_title("Top overall-label associations with paper I-L outcomes", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("")
    ax.set_ylabel("")
    fig.tight_layout()
    fig.savefig(TOP_OUTPUT, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def main():
    merged = build_merged()
    merged.to_csv(MERGED_CSV, index=False, encoding="utf-8-sig")

    assoc = association_tests(merged)
    assoc.to_csv(ASSOC_CSV, index=False, encoding="utf-8-sig")

    rates = group_rates(merged)
    rates["positive_rate"] = rates["positive_rate"].round(4)
    rates.to_csv(GROUP_RATE_CSV, index=False, encoding="utf-8-sig")

    bivar = bivariate_overall_tests(merged)
    bivar.to_csv(BIVARIATE_CSV, index=False, encoding="utf-8-sig")

    l1 = l1_logistic_models(merged)
    l1["coefficient"] = l1["coefficient"].round(4)
    l1["odds_ratio"] = l1["odds_ratio"].round(4)
    l1.to_csv(L1_CSV, index=False, encoding="utf-8-sig")

    plot_group_rate_heatmap(rates)
    plot_top_associations(bivar)

    print(f"merged_csv={MERGED_CSV}")
    print(f"assoc_csv={ASSOC_CSV}")
    print(f"rates_csv={GROUP_RATE_CSV}")
    print(f"bivariate_csv={BIVARIATE_CSV}")
    print(f"l1_csv={L1_CSV}")
    print(f"heatmap={HEATMAP_OUTPUT}")
    print(f"top_assoc={TOP_OUTPUT}")
    print(f"paper_n={len(merged)}")
    print(f"matched_overall_n={int(merged['matched_overall'].sum())}")


if __name__ == "__main__":
    main()
