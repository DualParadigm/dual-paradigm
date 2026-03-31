from __future__ import annotations

import csv
import math
from dataclasses import dataclass
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


WORKBOOK_PATH = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/新维度语料库0310_fixed (1).xlsx"
)
FEATURE_CSV_PATH = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/kappa_feature_summary.csv"
)
DIMENSION_CSV_PATH = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/kappa_dimension_summary.csv"
)
OVERALL_CSV_PATH = Path(
    "/Users/primav/Documents/GitHub/dual-paradigm/data/kappa_overall_summary.csv"
)


@dataclass
class KappaStats:
    n: int
    both_zero: int
    a_zero_b_one: int
    a_one_b_zero: int
    both_one: int
    coder_a_positive_rate: float
    coder_b_positive_rate: float
    observed_agreement: float
    expected_agreement: float
    kappa: float | None


def compute_binary_kappa(a_values: list[int], b_values: list[int]) -> KappaStats:
    n = len(a_values)
    both_zero = sum(1 for a, b in zip(a_values, b_values) if a == 0 and b == 0)
    a_zero_b_one = sum(1 for a, b in zip(a_values, b_values) if a == 0 and b == 1)
    a_one_b_zero = sum(1 for a, b in zip(a_values, b_values) if a == 1 and b == 0)
    both_one = sum(1 for a, b in zip(a_values, b_values) if a == 1 and b == 1)

    a_pos = (a_one_b_zero + both_one) / n
    b_pos = (a_zero_b_one + both_one) / n
    observed = (both_zero + both_one) / n
    expected = (a_pos * b_pos) + ((1 - a_pos) * (1 - b_pos))

    if math.isclose(1 - expected, 0.0, abs_tol=1e-12):
        kappa = None
    else:
        kappa = (observed - expected) / (1 - expected)

    return KappaStats(
        n=n,
        both_zero=both_zero,
        a_zero_b_one=a_zero_b_one,
        a_one_b_zero=a_one_b_zero,
        both_one=both_one,
        coder_a_positive_rate=a_pos,
        coder_b_positive_rate=b_pos,
        observed_agreement=observed,
        expected_agreement=expected,
        kappa=kappa,
    )


def interpretation(kappa: float | None) -> str:
    if kappa is None:
        return "undefined"
    if kappa < 0:
        return "poor"
    if kappa < 0.21:
        return "slight"
    if kappa < 0.41:
        return "fair"
    if kappa < 0.61:
        return "moderate"
    if kappa < 0.81:
        return "substantial"
    return "almost perfect"


def round_or_blank(value: float | None, digits: int = 4) -> str:
    if value is None:
        return ""
    return f"{value:.{digits}f}"


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["overall_matrix"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    data_rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1)
    ]

    feature_pairs: list[tuple[int, int, str, str]] = []
    for col in range(5, ws.max_column + 1, 2):
        a_header = str(headers[col - 1])
        b_header = str(headers[col])
        if not a_header.endswith("__coderA") or not b_header.endswith("__coderB"):
            raise ValueError(f"Unexpected matrix header pair: {a_header} / {b_header}")
        prefix = a_header[: -len("__coderA")]
        dimension, label = prefix.split("__", 1)
        feature_pairs.append((col, col + 1, dimension, label))

    feature_rows: list[dict[str, str]] = []
    dimension_bucket: dict[str, list[KappaStats]] = {}
    dimension_a: dict[str, list[int]] = {}
    dimension_b: dict[str, list[int]] = {}

    for a_col, b_col, dimension, label in feature_pairs:
        a_values = [int(row[a_col - 1] or 0) for row in data_rows]
        b_values = [int(row[b_col - 1] or 0) for row in data_rows]
        stats = compute_binary_kappa(a_values, b_values)

        feature_rows.append(
            {
                "dimension": dimension,
                "label": label,
                "n_cases": str(stats.n),
                "both_0": str(stats.both_zero),
                "a_0_b_1": str(stats.a_zero_b_one),
                "a_1_b_0": str(stats.a_one_b_zero),
                "both_1": str(stats.both_one),
                "coderA_positive_rate": round_or_blank(stats.coder_a_positive_rate),
                "coderB_positive_rate": round_or_blank(stats.coder_b_positive_rate),
                "observed_agreement": round_or_blank(stats.observed_agreement),
                "expected_agreement": round_or_blank(stats.expected_agreement),
                "kappa": round_or_blank(stats.kappa),
                "interpretation": interpretation(stats.kappa),
            }
        )

        dimension_bucket.setdefault(dimension, []).append(stats)
        dimension_a.setdefault(dimension, []).extend(a_values)
        dimension_b.setdefault(dimension, []).extend(b_values)

    dimension_rows: list[dict[str, str]] = []
    for dimension, stats_list in dimension_bucket.items():
        pooled_stats = compute_binary_kappa(dimension_a[dimension], dimension_b[dimension])
        kappas = [s.kappa for s in stats_list if s.kappa is not None]
        dimension_rows.append(
            {
                "dimension": dimension,
                "n_labels": str(len(stats_list)),
                "n_binary_decisions": str(len(dimension_a[dimension])),
                "mean_kappa": round_or_blank(mean(kappas) if kappas else None),
                "median_like_note": "feature mean across labels",
                "pooled_observed_agreement": round_or_blank(pooled_stats.observed_agreement),
                "pooled_expected_agreement": round_or_blank(pooled_stats.expected_agreement),
                "pooled_kappa": round_or_blank(pooled_stats.kappa),
                "pooled_interpretation": interpretation(pooled_stats.kappa),
            }
        )

    all_a: list[int] = []
    all_b: list[int] = []
    for dimension in dimension_a:
        all_a.extend(dimension_a[dimension])
        all_b.extend(dimension_b[dimension])
    overall_stats = compute_binary_kappa(all_a, all_b)
    overall_kappas = [
        float(row["kappa"]) for row in feature_rows if row["kappa"] != ""
    ]
    overall_rows = [
        {
            "n_cases": str(len(data_rows)),
            "n_labels": str(len(feature_pairs)),
            "n_binary_decisions": str(len(all_a)),
            "mean_feature_kappa": round_or_blank(mean(overall_kappas) if overall_kappas else None),
            "pooled_observed_agreement": round_or_blank(overall_stats.observed_agreement),
            "pooled_expected_agreement": round_or_blank(overall_stats.expected_agreement),
            "pooled_kappa": round_or_blank(overall_stats.kappa),
            "pooled_interpretation": interpretation(overall_stats.kappa),
        }
    ]

    with FEATURE_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(feature_rows[0].keys()))
        writer.writeheader()
        writer.writerows(feature_rows)

    with DIMENSION_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(dimension_rows[0].keys()))
        writer.writeheader()
        writer.writerows(dimension_rows)

    with OVERALL_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(overall_rows[0].keys()))
        writer.writeheader()
        writer.writerows(overall_rows)

    print(f"feature_csv={FEATURE_CSV_PATH}")
    print(f"dimension_csv={DIMENSION_CSV_PATH}")
    print(f"overall_csv={OVERALL_CSV_PATH}")
    print(f"n_cases={len(data_rows)}")
    print(f"n_labels={len(feature_pairs)}")
    print(f"overall_pooled_kappa={round_or_blank(overall_stats.kappa)}")


if __name__ == "__main__":
    main()
